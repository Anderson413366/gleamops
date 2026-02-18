#!/usr/bin/env python3
"""
Fix job_tasks: delete all, re-import with integer qc_weight,
then patch decimal values individually.
"""

import openpyxl
import json
import urllib.request
import urllib.error
import uuid
import re
import os
import sys
from datetime import datetime, date

# ── Config ────────────────────────────────────────────────────────────────────
SUPABASE_URL = os.environ.get('SUPABASE_URL', '')
SERVICE_KEY = os.environ.get('SUPABASE_SERVICE_ROLE_KEY', '')
TENANT_ID = os.environ.get('TENANT_ID', 'a0000000-0000-0000-0000-000000000001')
EXCEL_PATH = os.environ.get('EXCEL_PATH', './spreadsheets/Anderson_Cleaning_Database_UPDATED_Feb2026.xlsx')

if not SUPABASE_URL or not SERVICE_KEY:
    print('ERROR: Set SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY environment variables')
    print('  export SUPABASE_URL="https://your-project.supabase.co"')
    print('  export SUPABASE_SERVICE_ROLE_KEY="your-service-role-key"')
    sys.exit(1)

if not os.path.isfile(EXCEL_PATH):
    print(f'ERROR: Excel file not found at: {EXCEL_PATH}')
    print('  export EXCEL_PATH="/absolute/path/to/Anderson_Cleaning_Database_UPDATED_Feb2026.xlsx"')
    sys.exit(1)

HEADERS = {
    'apikey': SERVICE_KEY,
    'Authorization': f'Bearer {SERVICE_KEY}',
    'Content-Type': 'application/json',
    'Prefer': 'return=minimal',
}

def gen_uuid():
    return str(uuid.uuid4())

def clean_str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None

def clean_num(v, default=None):
    if v is None:
        return default
    try:
        return float(v)
    except (ValueError, TypeError):
        return default

def clean_bool(v, default=False):
    if v is None:
        return default
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    return s in ('true', '1', 'yes', 'y')

def read_sheet(wb, name):
    if name not in wb.sheetnames:
        print(f'  WARNING: Sheet "{name}" not found')
        return []
    ws = wb[name]
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = {}
        for i, h in enumerate(headers):
            if h and i < len(row):
                d[h] = row[i]
        if any(v is not None for v in d.values()):
            rows.append(d)
    return rows

def api_delete(table):
    url = f'{SUPABASE_URL}/rest/v1/{table}?tenant_id=eq.{TENANT_ID}'
    req = urllib.request.Request(url, method='DELETE', headers=HEADERS)
    try:
        urllib.request.urlopen(req)
    except urllib.error.HTTPError as e:
        print(f'  Delete {table} error: {e.read().decode()[:100]}')

def batch_insert(table, rows, batch_size=100):
    if not rows:
        print(f'  {table}: 0 rows, skipping')
        return
    total = len(rows)
    inserted = 0
    for i in range(0, total, batch_size):
        batch = rows[i:i+batch_size]
        data = json.dumps(batch, default=str).encode()
        req = urllib.request.Request(
            f'{SUPABASE_URL}/rest/v1/{table}',
            data=data,
            headers={**HEADERS, 'Prefer': 'return=minimal'},
        )
        try:
            urllib.request.urlopen(req)
            inserted += len(batch)
        except urllib.error.HTTPError as e:
            body = e.read().decode()[:200]
            print(f'  ERROR batch {i//batch_size}: {body}')
            # Retry individual rows
            for row in batch:
                single = json.dumps(row, default=str).encode()
                req2 = urllib.request.Request(
                    f'{SUPABASE_URL}/rest/v1/{table}',
                    data=single,
                    headers={**HEADERS, 'Prefer': 'return=minimal'},
                )
                try:
                    urllib.request.urlopen(req2)
                    inserted += 1
                except urllib.error.HTTPError as e2:
                    body2 = e2.read().decode()[:150]
                    tc = row.get('task_code', '?')
                    print(f'    SKIP {tc}: {body2}')
    print(f'  {table}: {inserted}/{total} rows inserted')

def main():
    print('Loading Excel...')
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    # First, get existing IDs for jobs and tasks from Supabase
    print('Fetching existing job IDs...')
    job_map = {}  # job_code → id
    offset = 0
    while True:
        url = f'{SUPABASE_URL}/rest/v1/site_jobs?select=id,job_code&tenant_id=eq.{TENANT_ID}&limit=1000&offset={offset}'
        req = urllib.request.Request(url, headers={
            'apikey': SERVICE_KEY,
            'Authorization': f'Bearer {SERVICE_KEY}',
        })
        resp = urllib.request.urlopen(req)
        data = json.loads(resp.read().decode())
        if not data:
            break
        for d in data:
            job_map[d['job_code']] = d['id']
        offset += 1000
    print(f'  Found {len(job_map)} jobs')

    print('Fetching existing task IDs...')
    task_map = {}  # task_code → id
    offset = 0
    while True:
        url = f'{SUPABASE_URL}/rest/v1/tasks?select=id,task_code&tenant_id=eq.{TENANT_ID}&limit=1000&offset={offset}'
        req = urllib.request.Request(url, headers={
            'apikey': SERVICE_KEY,
            'Authorization': f'Bearer {SERVICE_KEY}',
        })
        resp = urllib.request.urlopen(req)
        data = json.loads(resp.read().decode())
        if not data:
            break
        for d in data:
            task_map[d['task_code']] = d['id']
        offset += 1000
    print(f'  Found {len(task_map)} tasks')

    # Delete existing job_tasks
    print('Deleting existing job_tasks...')
    api_delete('job_tasks')
    print('  Deleted.')

    # Read Excel job tasks
    print('Reading Job Task sheet...')
    jt_rows = read_sheet(wb, 'Job Task')

    job_tasks = []
    decimal_patches = []  # (id, qc_weight) for later PATCH

    for r in jt_rows:
        job_code = clean_str(r.get('Job Code'))
        task_code = clean_str(r.get('Task Code'))
        if not job_code or not task_code:
            continue
        job_id = job_map.get(job_code)
        task_id = task_map.get(task_code)
        if not job_id or not task_id:
            continue

        raw_qc = clean_num(r.get('Qc Weight'), 0)
        int_qc = int(round(raw_qc))
        has_decimal = abs(raw_qc - int_qc) > 0.01

        row_id = gen_uuid()

        job_tasks.append({
            'id': row_id,
            'tenant_id': TENANT_ID,
            'job_id': job_id,
            'task_id': task_id,
            'task_code': task_code,
            'planned_minutes': int(round(clean_num(r.get('Planned Minutes'), 0))),
            'qc_weight': int_qc,
            'is_required': clean_bool(r.get('Is Required')),
            'status': clean_str(r.get('Status')) or 'ACTIVE',
            'notes': clean_str(r.get('Notes')),
        })

        if has_decimal:
            decimal_patches.append((row_id, raw_qc))

    print(f'  Prepared {len(job_tasks)} job_tasks ({len(decimal_patches)} need decimal patch)')

    # Insert all job_tasks with integer qc_weight
    print('Inserting job_tasks...')
    batch_insert('job_tasks', job_tasks)

    # Now patch the decimal values individually
    if decimal_patches:
        print(f'Patching {len(decimal_patches)} decimal qc_weight values...')
        patched = 0
        for row_id, qc_val in decimal_patches:
            url = f'{SUPABASE_URL}/rest/v1/job_tasks?id=eq.{row_id}'
            data = json.dumps({'qc_weight': qc_val}).encode()
            req = urllib.request.Request(url, data=data, method='PATCH', headers=HEADERS)
            try:
                urllib.request.urlopen(req)
                patched += 1
            except urllib.error.HTTPError as e:
                body = e.read().decode()[:100]
                print(f'  PATCH failed for {row_id}: {body}')
        print(f'  Patched {patched}/{len(decimal_patches)} rows')

    print('\nDone!')

if __name__ == '__main__':
    main()
