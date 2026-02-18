#!/usr/bin/env python3
"""
Import data from Anderson_Cleaning_Database_UPDATED_Feb2026.xlsx into GleamOps Supabase.
Comprehensive import mapping EVERY Excel column to its Supabase equivalent.

Steps:
  1. Delete all existing data (respecting FK order)
  2. Import real data from Excel (respecting FK order)
  3. Update system_sequences with max codes
"""

import openpyxl
import json
import urllib.request
import urllib.error
import uuid
import re
import sys
import os
from datetime import datetime, date, time as dtime

# ── Config ────────────────────────────────────────────────────────────────────
SUPABASE_URL = os.environ.get('SUPABASE_URL', '')
SERVICE_KEY = os.environ.get('SUPABASE_SERVICE_ROLE_KEY', '')
TENANT_ID = os.environ.get('TENANT_ID', 'a0000000-0000-0000-0000-000000000001')
EXCEL_PATH = os.environ.get('EXCEL_PATH', './spreadsheets/Anderson_Cleaning_Database_UPDATED_Feb2026.xlsx')

if not SUPABASE_URL or not SERVICE_KEY:
    print('ERROR: Set SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY environment variables')
    print('  export SUPABASE_URL="https://your-project.supabase.co"')
    print('  export SUPABASE_SERVICE_ROLE_KEY="your-service-role-key"')
    sys.exit(1)

if not os.path.isfile(EXCEL_PATH):
    print(f'ERROR: Excel file not found at: {EXCEL_PATH}')
    print('  export EXCEL_PATH="/absolute/path/to/Anderson_Cleaning_Database_UPDATED_Feb2026.xlsx"')
    sys.exit(1)

HEADERS = {
    'apikey': SERVICE_KEY,
    'Authorization': f'Bearer {SERVICE_KEY}',
    'Content-Type': 'application/json',
    'Prefer': 'return=minimal',
}

# ── Helpers ───────────────────────────────────────────────────────────────────
def gen_uuid():
    return str(uuid.uuid4())

def clean_str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s and s not in ('N/A', 'n/a', 'None', 'NULL', '-') else None

def clean_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, date):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    if not s or s in ('N/A', 'n/a', 'None', '', '-'):
        return None
    for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%m/%d/%y', '%Y-%m-%dT%H:%M:%S'):
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return None

def clean_time(v):
    """Convert to HH:MM:SS time string."""
    if v is None:
        return None
    if isinstance(v, dtime):
        return v.strftime('%H:%M:%S')
    if isinstance(v, datetime):
        return v.strftime('%H:%M:%S')
    s = str(v).strip()
    if not s or s in ('N/A', 'n/a', 'None', '', '-'):
        return None
    # Try parsing common formats
    for fmt in ('%H:%M:%S', '%H:%M', '%I:%M %p', '%I:%M%p', '%I:%M:%S %p'):
        try:
            return datetime.strptime(s, fmt).strftime('%H:%M:%S')
        except ValueError:
            continue
    return None

def clean_num(v, default=None):
    if v is None:
        return default
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip().replace(',', '').replace('$', '').replace('%', '')
    if not s or s in ('N/A', 'n/a', 'None', '-', ''):
        return default
    try:
        return float(s)
    except ValueError:
        return default

def clean_int(v, default=None):
    n = clean_num(v, default)
    if n is None:
        return default
    return int(round(n))

def clean_bool(v, default=True):
    if v is None:
        return default
    if isinstance(v, bool):
        return v
    s = str(v).strip().upper()
    if s in ('TRUE', 'YES', '1', 'Y', 'ACTIVE'):
        return True
    if s in ('FALSE', 'NO', '0', 'N', 'INACTIVE'):
        return False
    return default

def clean_role(raw):
    """Clean staff role, removing emoji prefixes and level suffixes."""
    if not raw:
        return 'CLEANER'
    role_clean = re.sub(r'^[^\w]+', '', raw).strip()
    role_clean = role_clean.split('•')[0].strip()
    role_clean = role_clean.split('·')[0].strip()
    role_map = {
        'Owner': 'OWNER_ADMIN', 'Manager': 'MANAGER', 'Supervisor': 'SUPERVISOR',
        'Cleaner': 'CLEANER', 'Inspector': 'INSPECTOR', 'Sales': 'SALES',
        'Admin': 'OWNER_ADMIN', 'Lead': 'SUPERVISOR', 'Account Manager': 'MANAGER',
        'Operations Manager': 'MANAGER', 'Project Manager': 'MANAGER',
    }
    return role_map.get(role_clean, 'CLEANER')

def map_status(raw, mapping, default='ACTIVE'):
    if not raw:
        return default
    s = str(raw).strip()
    if s in mapping:
        return mapping[s]
    return s.upper().replace(' ', '_')

def read_sheet(wb, sheet_name):
    if sheet_name not in wb.sheetnames:
        print(f'  Sheet "{sheet_name}" not found')
        return []
    ws = wb[sheet_name]
    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v:
            headers.append(str(v).strip())
        else:
            headers.append(f'_col{c}')
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {}
        has_data = False
        for c, h in enumerate(headers, 1):
            v = ws.cell(r, c).value
            if v is not None:
                has_data = True
            row[h] = v
        if has_data:
            rows.append(row)
    return rows

def delete_all(table):
    url = f'{SUPABASE_URL}/rest/v1/{table}?id=not.is.null'
    req = urllib.request.Request(url, headers=HEADERS, method='DELETE')
    try:
        urllib.request.urlopen(req)
        print(f'  Deleted from {table}')
    except urllib.error.HTTPError as e:
        error_body = e.read().decode()
        if '404' in str(e.code):
            print(f'  {table}: table not found, skipping')
        else:
            print(f'  ERROR deleting {table}: {e.code} - {error_body[:200]}')

def batch_insert(table, rows, batch_size=100):
    if not rows:
        print(f'  {table}: 0 rows, skipping')
        return 0
    total = 0
    for i in range(0, len(rows), batch_size):
        batch = rows[i:i+batch_size]
        url = f'{SUPABASE_URL}/rest/v1/{table}'
        body = json.dumps(batch, default=str).encode()
        h = dict(HEADERS)
        h['Prefer'] = 'return=minimal,resolution=ignore-duplicates'
        req = urllib.request.Request(url, data=body, headers=h, method='POST')
        try:
            urllib.request.urlopen(req)
            total += len(batch)
        except urllib.error.HTTPError as e:
            error_body = e.read().decode()
            print(f'  ERROR inserting {table} batch {i//batch_size}: {error_body[:400]}')
            for row in batch:
                try:
                    req2 = urllib.request.Request(url, data=json.dumps(row, default=str).encode(), headers=h, method='POST')
                    urllib.request.urlopen(req2)
                    total += 1
                except urllib.error.HTTPError as e2:
                    err = e2.read().decode()
                    code_val = row.get(next((k for k in row if 'code' in k.lower()), 'id'), '?')
                    print(f'    SKIP {table} row {code_val}: {err[:200]}')
    print(f'  {table}: {total}/{len(rows)} rows inserted')
    return total


# ── Step 1: Delete all existing data ─────────────────────────────────────────
def delete_all_data():
    print('\n=== STEP 1: Deleting all existing data ===\n')
    delete_order = [
        # Sales pipeline children
        'sales_followup_sends', 'sales_followup_sequences',
        'sales_email_events', 'sales_proposal_sends',
        'sales_proposal_marketing_inserts', 'sales_proposal_attachments',
        'sales_proposal_signatures', 'sales_proposal_pricing_options',
        'sales_proposals',
        'sales_bid_pricing_results', 'sales_bid_workload_results',
        'sales_bid_burden', 'sales_bid_labor_rates', 'sales_bid_schedule',
        'sales_bid_area_tasks', 'sales_bid_areas',
        'sales_bid_sites', 'sales_bid_general_tasks',
        'sales_bid_consumables', 'sales_bid_supply_allowances',
        'sales_bid_supply_kits', 'sales_bid_equipment_plan_items',
        'sales_bid_overhead', 'sales_bid_pricing_strategy',
        'sales_bid_versions', 'sales_bids',
        'sales_opportunities', 'sales_prospect_contacts', 'sales_prospects',
        'sales_marketing_inserts', 'sales_followup_templates',
        'sales_production_rates',
        # Conversion / Operations
        'sales_conversion_events', 'sales_bid_conversions',
        'ticket_asset_checkouts', 'site_asset_requirements',
        'ticket_photos', 'ticket_checklist_items', 'ticket_checklists',
        'checklist_template_items', 'checklist_templates',
        'ticket_assignments', 'work_tickets', 'recurrence_rules',
        # Inspections
        'inspection_issues', 'inspection_items', 'inspections',
        'inspection_template_items', 'inspection_templates',
        # Timekeeping
        'timesheet_approvals', 'timesheets',
        'time_exceptions', 'time_entries', 'time_events',
        'alerts', 'geofences',
        # Training / Safety
        'training_completions', 'training_courses',
        'safety_documents', 'key_event_log',
        'vehicle_checkouts', 'pay_rate_history', 'staff_certifications',
        'user_access_grants', 'user_team_memberships',
        # Inventory / Assets
        'inventory_count_details', 'inventory_counts',
        'supply_kit_items', 'supply_kits',
        'supply_orders', 'vehicle_maintenance',
        'equipment_assignments', 'equipment',
        'key_inventory', 'vehicles',
        'site_supplies', 'supply_catalog',
        # Staff / Jobs
        'job_staff_assignments', 'job_tasks', 'job_logs', 'site_jobs',
        'subcontractors', 'staff_positions',
        # CRM
        'timeline_events', 'contacts',
        'service_tasks', 'task_production_rates', 'tasks', 'services',
        'sites', 'clients', 'staff',
        # User/RBAC
        'user_profiles', 'user_client_access',
        # System
        'audit_events', 'notifications', 'files',
    ]
    for table in delete_order:
        delete_all(table)
    delete_all('lookups')
    delete_all('status_transitions')
    print('\n  All existing data deleted.')


# ── Step 2: Import Excel data ────────────────────────────────────────────────
def import_data():
    print('\n=== STEP 2: Importing Excel data ===\n')
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # ID maps: Excel code → Supabase UUID
    client_ids = {}
    site_ids = {}
    staff_ids = {}
    service_ids = {}
    task_ids = {}
    job_ids = {}
    supply_ids = {}
    position_ids = {}
    equipment_ids = {}
    subcontractor_ids = {}
    count_ids = {}

    CLIENT_STATUS = {'Active': 'ACTIVE', 'Inactive': 'INACTIVE', 'On Hold': 'ON_HOLD',
                     'Prospect': 'PROSPECT', 'Cancelled': 'CANCELED', 'Canceled': 'CANCELED'}
    SITE_STATUS = {'Active': 'ACTIVE', 'Inactive': 'INACTIVE', 'On Hold': 'ON_HOLD',
                   'Canceled': 'CANCELED', 'Cancelled': 'CANCELED'}
    STAFF_STATUS = {'Active': 'ACTIVE', 'Inactive': 'INACTIVE', 'On Leave': 'ON_LEAVE',
                    'Terminated': 'TERMINATED'}
    JOB_STATUS = {'Active': 'ACTIVE', 'Inactive': 'INACTIVE', 'On Hold': 'ON_HOLD',
                  'Canceled': 'CANCELED', 'Cancelled': 'CANCELED', 'Completed': 'COMPLETED'}
    FREQ_MAP = {
        'Daily': 'DAILY', 'Weekly': 'WEEKLY', 'Monthly': 'MONTHLY',
        'Bi-Weekly': 'BIWEEKLY', 'Biweekly': 'BIWEEKLY',
        '2x Weekly': '2X_WEEK', '3x Weekly': '3X_WEEK', '4x Weekly': '4X_WEEK', '5x Weekly': '5X_WEEK',
        'As Needed': 'AS_NEEDED', 'One-Time': 'AS_NEEDED',
    }

    # ── 2a. Lookups ──────────────────────────────────────────────────────
    print('Importing lookups...')
    lookup_rows = read_sheet(wb, 'Lookups')
    lookups = []
    for r in lookup_rows:
        cat = clean_str(r.get('Category'))
        code = clean_str(r.get('Code'))
        label = clean_str(r.get('Value'))
        if not cat or not code or not label:
            continue
        lookups.append({
            'id': gen_uuid(),
            'tenant_id': None,
            'category': cat,
            'code': code,
            'label': label,
            'sort_order': clean_int(r.get('Sort'), 0),
            'is_active': clean_bool(r.get('Active')),
        })
    batch_insert('lookups', lookups)

    # ── 2b. Staff Positions ──────────────────────────────────────────────
    print('Importing staff positions...')
    pos_rows = read_sheet(wb, 'Staff Position')
    positions = []
    for r in pos_rows:
        code = clean_str(r.get('Position Code'))
        title = clean_str(r.get('Position Name'))
        if not code or not title:
            continue
        pid = gen_uuid()
        position_ids[code] = pid
        positions.append({
            'id': pid,
            'tenant_id': TENANT_ID,
            'position_code': code,
            'title': title,
            'pay_grade': clean_str(r.get('Skill Level')),
            'notes': clean_str(r.get('Notes')),
            'is_active': clean_bool(r.get('Is Active')),
        })
    batch_insert('staff_positions', positions)

    # ── 2c. Services ─────────────────────────────────────────────────────
    print('Importing services...')
    svc_rows = read_sheet(wb, 'Service')
    services = []
    for r in svc_rows:
        code = clean_str(r.get('Service Code'))
        name = clean_str(r.get('Service Name'))
        if not code or not name:
            continue
        sid = gen_uuid()
        service_ids[code] = sid
        services.append({
            'id': sid,
            'tenant_id': TENANT_ID,
            'service_code': code,
            'name': name,
            'description': clean_str(r.get('Description')),
        })
    batch_insert('services', services)

    # ── 2d. Tasks ────────────────────────────────────────────────────────
    print('Importing tasks...')
    task_rows = read_sheet(wb, 'Task')
    tasks = []
    for r in task_rows:
        code = clean_str(r.get('Task Code'))
        name = clean_str(r.get('Task Name'))
        if not code or not name:
            continue
        tid = gen_uuid()
        task_ids[code] = tid

        freq_raw = clean_str(r.get('Frequency')) or 'DAILY'
        freq = FREQ_MAP.get(freq_raw, freq_raw.upper().replace(' ', '_').replace('-', '_'))

        tasks.append({
            'id': tid,
            'tenant_id': TENANT_ID,
            'task_code': code,
            'name': name,
            'category': clean_str(r.get('Category')),
            'subcategory': clean_str(r.get('Subcategory')),
            'area_type': clean_str(r.get('Area Type')),
            'floor_type': clean_str(r.get('Floor Type')),
            'priority_level': clean_str(r.get('Priority Level')),
            'default_minutes': clean_int(r.get('Default Minutes')),
            'production_rate_sqft_per_hour': clean_num(r.get('Production Rate')),
            'unit_code': clean_str(r.get('Default UOM')) or 'SQFT_1000',
            'spec_description': clean_str(r.get('Spec Description')),
            'work_description': clean_str(r.get('Work Description')),
            'tools_materials': clean_str(r.get('Tools Materials')),
            'notes': clean_str(r.get('Notes')),
            'is_active': clean_bool(r.get('Is Active')),
        })
    batch_insert('tasks', tasks)

    # ── 2e. Service Tasks ────────────────────────────────────────────────
    print('Importing service tasks...')
    st_rows = read_sheet(wb, 'Service Task')
    service_tasks = []
    seen_st = set()
    for r in st_rows:
        svc_code = clean_str(r.get('Service Code'))
        tsk_code = clean_str(r.get('Task Code'))
        if not svc_code or not tsk_code:
            continue
        sid = service_ids.get(svc_code)
        tid = task_ids.get(tsk_code)
        if not sid or not tid:
            continue
        key = (sid, tid)
        if key in seen_st:
            continue
        seen_st.add(key)

        freq_raw = clean_str(r.get('Typical Frequency')) or 'DAILY'
        freq = FREQ_MAP.get(freq_raw, freq_raw.upper().replace(' ', '_').replace('-', '_'))

        service_tasks.append({
            'id': gen_uuid(),
            'tenant_id': TENANT_ID,
            'service_id': sid,
            'task_id': tid,
            'frequency_default': freq,
            'sequence_order': clean_int(r.get('Sequence Order'), 0),
            'priority_level': clean_str(r.get('Priority Level')),
            'is_required': clean_bool(r.get('Is Required')),
            'estimated_minutes': clean_int(r.get('Estimated Minutes')),
            'quality_weight': clean_num(r.get('Quality Weight'), 1),
            'notes': clean_str(r.get('Notes')),
        })
    batch_insert('service_tasks', service_tasks)

    # ── 2f. Clients ──────────────────────────────────────────────────────
    print('Importing clients...')
    cli_rows = read_sheet(wb, 'Client')
    clients = []
    for r in cli_rows:
        code = clean_str(r.get('Client Code'))
        name = clean_str(r.get('Client Name'))
        if not code or not name:
            continue
        cid = gen_uuid()
        client_ids[code] = cid

        status = map_status(r.get('Client Status'), CLIENT_STATUS)

        billing_addr = {}
        street = clean_str(r.get('Billing Address'))
        if street:
            billing_addr = {
                'street': street,
                'suite': clean_str(r.get('Suite/Unit')),
                'city': clean_str(r.get('Billing City')),
                'state': clean_str(r.get('Billing State')),
                'zip': clean_str(r.get('Billing Zip')),
            }

        clients.append({
            'id': cid,
            'tenant_id': TENANT_ID,
            'client_code': code,
            'name': name,
            'status': status,
            'billing_address': billing_addr if billing_addr else None,
            'client_since': clean_date(r.get('Client Since')),
            'client_type': clean_str(r.get('Client Type')),
            'industry': clean_str(r.get('Industry')),
            'bill_to_name': clean_str(r.get('Bill To Name')),
            'payment_terms': clean_str(r.get('Payment Terms')),
            'po_required': clean_bool(r.get('PO Required'), False),
            'insurance_required': clean_bool(r.get('Insurance Required'), False),
            'insurance_expiry': clean_date(r.get('Insurance Expiry Date')),
            'credit_limit': clean_num(r.get('Credit Limit')),
            'website': clean_str(r.get('Website')),
            'tax_id': clean_str(r.get('Tax ID')),
            'contract_start_date': clean_date(r.get('Contract Start Date')),
            'contract_end_date': clean_date(r.get('Contract End Date')),
            'auto_renewal': clean_bool(r.get('Auto Renewal'), False),
            'invoice_frequency': clean_str(r.get('Invoice Frequency')),
            'notes': clean_str(r.get('Notes')),
        })
    batch_insert('clients', clients)

    # ── 2g. Staff ────────────────────────────────────────────────────────
    # The Excel has duplicate rows per person: -A suffix (basic data) and
    # -B suffix (complete data with pay, contact, updated status). We keep
    # only the -B row (or whichever is more complete), strip the suffix,
    # and store a clean STF-NNNN code.
    print('Importing staff...')
    stf_rows = read_sheet(wb, 'Staff')

    def strip_staff_suffix(code):
        """STF-1001-A → STF-1001, STF-1001-B → STF-1001"""
        if not code:
            return code
        return re.sub(r'-[AB]$', '', code)

    # Group rows by base code, prefer -B rows (more complete data)
    staff_by_base = {}  # base_code → (priority, row)
    for r in stf_rows:
        raw_code = clean_str(r.get('Staff Code'))
        if not raw_code:
            continue
        first = clean_str(r.get('First Name'))
        last = clean_str(r.get('Last Name'))
        # Skip junk rows (header echoes)
        if first and first.lower() == 'first name':
            continue

        base_code = strip_staff_suffix(raw_code)
        # -B suffix gets priority 2 (preferred), -A gets 1, no suffix gets 3
        if raw_code.endswith('-B'):
            priority = 2
        elif raw_code.endswith('-A'):
            priority = 1
        else:
            priority = 3

        existing = staff_by_base.get(base_code)
        if not existing or priority > existing[0]:
            staff_by_base[base_code] = (priority, r, raw_code)

    dupes_skipped = len(stf_rows) - len(staff_by_base)
    if dupes_skipped > 0:
        print(f'  Deduped: {dupes_skipped} duplicate staff rows removed (kept -B variants)')

    staff_list = []
    staff_supervisor_map = {}  # base_code → supervisor base_code
    raw_to_base = {}  # raw_code → base_code (for FK resolution from other sheets)
    for base_code, (priority, r, raw_code) in staff_by_base.items():
        first = clean_str(r.get('First Name'))
        last = clean_str(r.get('Last Name'))
        full_name = f'{first or ""} {last or ""}'.strip()
        if not full_name:
            full_name = base_code

        sid = gen_uuid()
        staff_ids[base_code] = sid
        # Also map the original raw code (with suffix) to the same UUID
        staff_ids[raw_code] = sid
        raw_to_base[raw_code] = base_code

        role = clean_role(clean_str(r.get('Staff Role')))
        status = map_status(r.get('Staff Status'), STAFF_STATUS)

        address = {}
        street = clean_str(r.get('Street Address'))
        if street:
            address = {
                'street': street,
                'suite': clean_str(r.get('Suite/Unit')),
                'city': clean_str(r.get('City')),
                'state': clean_str(r.get('State')),
                'zip': clean_str(r.get('ZIP Code')),
            }

        sup_code = clean_str(r.get('Supervisor Code'))
        if sup_code:
            staff_supervisor_map[base_code] = strip_staff_suffix(sup_code)

        staff_list.append({
            'id': sid,
            'tenant_id': TENANT_ID,
            'staff_code': base_code,
            'full_name': full_name,
            'first_name': first,
            'last_name': last,
            'preferred_name': clean_str(r.get('Preferred Name')),
            'role': role,
            'staff_status': status,
            'staff_type': clean_str(r.get('Staff Type')),
            'employment_type': clean_str(r.get('Employment Type')),
            'hire_date': clean_date(r.get('Hire Date')),
            'termination_date': clean_date(r.get('Termination Date')),
            'email': clean_str(r.get('Email')),
            'phone': clean_str(r.get('Mobile Phone')),
            'mobile_phone': clean_str(r.get('Mobile Phone')),
            'pay_rate': clean_num(r.get('Pay Rate')),
            'schedule_type': clean_str(r.get('Schedule Type')),
            'address': address if address else None,
            'emergency_contact_name': clean_str(r.get('Emergency Contact Name')),
            'emergency_contact_phone': clean_str(r.get('Emergency Contact Phone')),
            'emergency_contact_relationship': clean_str(r.get('Emergency Contact Relationship')),
            'certifications': clean_str(r.get('Certifications')),
            'performance_rating': clean_num(r.get('Performance Rating')),
            'background_check_date': clean_date(r.get('Background Check Date')),
            'photo_url': clean_str(r.get('Photo URL')),
            'notes': clean_str(r.get('Notes')),
        })
    batch_insert('staff', staff_list)

    # Patch supervisor_id references (using base codes)
    if staff_supervisor_map:
        print(f'  Patching {len(staff_supervisor_map)} supervisor references...')
        patched = 0
        for base_code, sup_base_code in staff_supervisor_map.items():
            staff_id = staff_ids.get(base_code)
            sup_id = staff_ids.get(sup_base_code)
            if staff_id and sup_id:
                url = f'{SUPABASE_URL}/rest/v1/staff?id=eq.{staff_id}'
                data = json.dumps({'supervisor_id': sup_id}).encode()
                req = urllib.request.Request(url, data=data, method='PATCH', headers=HEADERS)
                try:
                    urllib.request.urlopen(req)
                    patched += 1
                except urllib.error.HTTPError:
                    pass
        print(f'  Patched {patched}/{len(staff_supervisor_map)} supervisors')

    # ── 2h. Sites ────────────────────────────────────────────────────────
    print('Importing sites...')
    site_rows = read_sheet(wb, 'Site')
    sites = []
    site_contact_refs = []  # (site_code, primary_code, emergency_code, supervisor_code) for later
    for r in site_rows:
        code = clean_str(r.get('Site Code'))
        name = clean_str(r.get('Site Name'))
        if not code or not name:
            continue

        client_code = clean_str(r.get('Client Code'))
        client_id = client_ids.get(client_code)
        if not client_id:
            print(f'    WARN: Site {code} has unknown client {client_code}, skipping')
            continue

        sid = gen_uuid()
        site_ids[code] = sid

        status = map_status(r.get('Site Status'), SITE_STATUS)

        address = {}
        street = clean_str(r.get('Street Address'))
        if street:
            address = {
                'street': street,
                'suite': clean_str(r.get('Suite/Unit')),
                'city': clean_str(r.get('City')),
                'state': clean_str(r.get('State')),
                'zip': clean_str(r.get('ZIP Code')),
            }

        sup_code = clean_str(r.get('Supervisor Code'))
        sup_id = staff_ids.get(sup_code) if sup_code else None

        sites.append({
            'id': sid,
            'tenant_id': TENANT_ID,
            'client_id': client_id,
            'site_code': code,
            'name': name,
            'status': status,
            'status_date': clean_date(r.get('Status Date')),
            'status_reason': clean_str(r.get('Status Reason')),
            'service_start_date': clean_date(r.get('Service Start Date')),
            'address': address if address else {},
            'alarm_code': clean_str(r.get('Alarm Code')),
            'alarm_system': clean_str(r.get('Alarm System')),
            'alarm_company': clean_str(r.get('Alarm Company')),
            'security_protocol': clean_str(r.get('Security Protocol')),
            'access_notes': clean_str(r.get('Entry Instructions')),
            'entry_instructions': clean_str(r.get('Entry Instructions')),
            'parking_instructions': clean_str(r.get('Parking Instructions')),
            'square_footage': clean_num(r.get('Total Cleanable SqFt')),
            'number_of_floors': clean_int(r.get('Number Of Floors')),
            'employees_on_site': clean_int(r.get('Employees On Site')),
            'earliest_start_time': clean_time(r.get('Earliest Start Time')),
            'latest_start_time': clean_time(r.get('Latest Start Time')),
            'business_hours_start': clean_time(r.get('Business Hours Start')),
            'business_hours_end': clean_time(r.get('Business Hours End')),
            'weekend_access': clean_bool(r.get('Weekend Access'), False),
            'janitorial_closet_location': clean_str(r.get('Janitorial Closet Location')),
            'supply_storage_location': clean_str(r.get('Supply Storage Location')),
            'water_source_location': clean_str(r.get('Water Source Location')),
            'dumpster_location': clean_str(r.get('Dumpster Location')),
            'supervisor_id': sup_id,
            'risk_level': clean_str(r.get('Risk Level')),
            'priority_level': clean_str(r.get('Priority Level')),
            'osha_compliance_required': clean_bool(r.get('OSHA Compliance Required'), False),
            'background_check_required': clean_bool(r.get('Background Check Required'), False),
            'last_inspection_date': clean_date(r.get('Last Inspection Date')),
            'next_inspection_date': clean_date(r.get('Next Inspection Date')),
            'notes': clean_str(r.get('Notes')),
        })
    batch_insert('sites', sites)

    # ── 2i. Subcontractors ───────────────────────────────────────────────
    print('Importing subcontractors...')
    sub_rows = read_sheet(wb, 'Subcontractor')
    subs = []
    for r in sub_rows:
        code = clean_str(r.get('Subcontractor Code'))
        name = clean_str(r.get('Subcontractor Name'))
        if not code or not name:
            continue

        sub_id = gen_uuid()
        subcontractor_ids[code] = sub_id

        address = {}
        street = clean_str(r.get('Street Address'))
        if street:
            address = {
                'street': street,
                'city': clean_str(r.get('City')),
                'state': clean_str(r.get('State')),
                'zip': clean_str(r.get('ZIP Code')),
            }

        subs.append({
            'id': sub_id,
            'tenant_id': TENANT_ID,
            'subcontractor_code': code,
            'company_name': name,
            'contact_name': clean_str(r.get('Contact Name')),
            'contact_title': clean_str(r.get('Contact Title')),
            'email': clean_str(r.get('Email')),
            'phone': clean_str(r.get('Business Phone')),
            'business_phone': clean_str(r.get('Business Phone')),
            'mobile_phone': clean_str(r.get('Mobile Phone')),
            'website': clean_str(r.get('Website')),
            'address': address if address else None,
            'status': 'ACTIVE',
            'services_provided': clean_str(r.get('Services Provided')),
            'license_number': clean_str(r.get('License Number')),
            'license_expiry': clean_date(r.get('License Expiry')),
            'insurance_company': clean_str(r.get('Insurance Company')),
            'insurance_policy_number': clean_str(r.get('Insurance Policy Number')),
            'insurance_expiry': clean_date(r.get('Insurance Expiry')),
            'hourly_rate': clean_num(r.get('Hourly Rate')),
            'payment_terms': clean_str(r.get('Payment Terms')),
            'tax_id': clean_str(r.get('Tax ID')),
            'w9_on_file': clean_bool(r.get('W9 On File'), False),
            'notes': clean_str(r.get('Notes')),
        })
    batch_insert('subcontractors', subs)

    # ── 2j. Site Jobs ────────────────────────────────────────────────────
    print('Importing site jobs...')
    job_rows = read_sheet(wb, 'Site Job')
    jobs = []
    seen_job_codes = set()
    for r in job_rows:
        code = clean_str(r.get('Job Code'))
        name = clean_str(r.get('Job Name'))
        site_code = clean_str(r.get('Site Code'))
        if not code or not name:
            continue
        if code in seen_job_codes:
            print(f'    WARN: Duplicate job code {code}, skipping')
            continue
        seen_job_codes.add(code)

        site_id = site_ids.get(site_code)
        if not site_id:
            print(f'    WARN: Job {code} has unknown site {site_code}, skipping')
            continue

        jid = gen_uuid()
        job_ids[code] = jid

        svc_code = clean_str(r.get('Service Code'))
        svc_id = service_ids.get(svc_code)

        status = map_status(r.get('Job Status'), JOB_STATUS)

        freq_raw = clean_str(r.get('Frequency')) or 'WEEKLY'
        freq = FREQ_MAP.get(freq_raw, freq_raw.upper().replace(' ', '_').replace('-', '_'))

        sub_code = clean_str(r.get('Subcontractor Code'))
        sub_id = subcontractor_ids.get(sub_code) if sub_code else None

        jobs.append({
            'id': jid,
            'tenant_id': TENANT_ID,
            'site_id': site_id,
            'job_code': code,
            'job_name': name,
            'status': status,
            'frequency': freq,
            'service_id': svc_id,
            'job_type': clean_str(r.get('Job Type')),
            'priority_level': clean_str(r.get('Priority Level')),
            'schedule_days': clean_str(r.get('Schedule Days')),
            'staff_needed': clean_int(r.get('Staff Needed')),
            'start_time': clean_time(r.get('Start Time')),
            'end_time': clean_time(r.get('End Time')),
            'estimated_hours_per_service': clean_num(r.get('Estimated Hours Svc')),
            'estimated_hours_per_month': clean_num(r.get('Estimated Hours Mo')),
            'last_service_date': clean_date(r.get('Last Service Date')),
            'next_service_date': clean_date(r.get('Next Service Date')),
            'quality_score': clean_num(r.get('Quality Score')),
            'billing_uom': clean_str(r.get('Billing UOM')) or 'MONTHLY',
            'billing_amount': clean_num(r.get('Billing Amount')),
            'job_assigned_to': clean_str(r.get('Job Assigned To')),
            'subcontractor_id': sub_id,
            'invoice_description': clean_str(r.get('Invoice Service Description')),
            'specifications': clean_str(r.get('Job Specifications')),
            'special_requirements': clean_str(r.get('Special Requirements')),
            'notes': clean_str(r.get('Notes')),
        })
    batch_insert('site_jobs', jobs)

    # ── 2k. Job Tasks ────────────────────────────────────────────────────
    # Deduplicate by (job_id, task_id) — keep last occurrence from Excel
    print('Importing job tasks...')
    jt_rows = read_sheet(wb, 'Job Task')
    jt_map = {}  # (job_id, task_id) → row dict
    jt_skipped = 0
    for r in jt_rows:
        job_code = clean_str(r.get('Job Code'))
        task_code = clean_str(r.get('Task Code'))
        if not job_code or not task_code:
            continue
        job_id = job_ids.get(job_code)
        task_id = task_ids.get(task_code)
        if not job_id or not task_id:
            continue

        key = (job_id, task_id)
        if key in jt_map:
            jt_skipped += 1
        jt_map[key] = {
            'id': gen_uuid(),
            'tenant_id': TENANT_ID,
            'job_id': job_id,
            'task_id': task_id,
            'task_code': task_code,
            'task_name': clean_str(r.get('Task Name')),
            'planned_minutes': clean_int(r.get('Planned Minutes'), 0),
            'qc_weight': clean_num(r.get('Qc Weight'), 1),
            'is_required': clean_bool(r.get('Is Required')),
            'status': clean_str(r.get('Status')) or 'ACTIVE',
            'notes': clean_str(r.get('Notes')),
        }
    job_tasks = list(jt_map.values())
    if jt_skipped:
        print(f'  Deduped: {jt_skipped} duplicate job+task combos removed')
    batch_insert('job_tasks', job_tasks)

    # ── 2l. Supply Catalog ───────────────────────────────────────────────
    print('Importing supplies...')
    sup_rows = read_sheet(wb, 'Supply')
    supplies = []
    for r in sup_rows:
        code = clean_str(r.get('\U0001f3f7\ufe0f Supply_Code'))
        name = clean_str(r.get('\U0001f1fa\U0001f1f8 Supply_Name_EN'))
        if not code or not name:
            continue
        supid = gen_uuid()
        supply_ids[code] = supid

        supplies.append({
            'id': supid,
            'tenant_id': TENANT_ID,
            'code': code,
            'name': name,
            'description': clean_str(r.get('\U0001f4dd Description_EN')),
            'category': clean_str(r.get('\U0001f4c1 Supply_Category')),
            'supply_status': clean_str(r.get('\U0001f504 Supply_Status')) or 'ACTIVE',
            'unit': clean_str(r.get('Unit_Of_Measure')) or 'EA',
            'pack_size': clean_str(r.get('Pack_Size')),
            'min_stock_level': clean_int(r.get('\u26a0\ufe0f Min_Stock_Level')),
            'brand': clean_str(r.get('Brand')),
            'manufacturer': clean_str(r.get('Manufacturer')),
            'model_number': clean_str(r.get('Model_Number')),
            'markup_percentage': clean_num(r.get('Markup_Percentage')),
            'billing_rate': clean_num(r.get('Billing_Rate')),
            'preferred_vendor': clean_str(r.get('Preferred_Vendor')),
            'vendor_sku': clean_str(r.get('Vendor_Item_Sku')),
            'eco_rating': clean_str(r.get('Eco_Rating')),
            'ppe_required': clean_bool(r.get('PPE'), False),
            'sds_url': clean_str(r.get('SDS_Link')),
            'image_url': clean_str(r.get('\U0001f5bc\ufe0f Supply_Image_URL')),
            'notes': clean_str(r.get('Notes')),
        })
    batch_insert('supply_catalog', supplies)

    # ── 2m. Equipment ────────────────────────────────────────────────────
    print('Importing equipment...')
    eq_rows = read_sheet(wb, 'Equipment')
    equip_list = []
    for r in eq_rows:
        code = clean_str(r.get('Equipment Code'))
        name = clean_str(r.get('Equipment Name'))
        if not code:
            continue
        if not name:
            name = code

        eid = gen_uuid()
        equipment_ids[code] = eid

        condition = clean_str(r.get('Condition')) or 'GOOD'
        condition = condition.upper().replace(' ', '_')

        equip_list.append({
            'id': eid,
            'tenant_id': TENANT_ID,
            'equipment_code': code,
            'name': name,
            'equipment_type': clean_str(r.get('Equipment Type')),
            'equipment_category': clean_str(r.get('Equipment Category')),
            'manufacturer': clean_str(r.get('Manufacturer')),
            'brand': clean_str(r.get('Brand')),
            'model_number': clean_str(r.get('Model Number')),
            'condition': condition,
            'serial_number': clean_str(r.get('Serial Number')),
            'purchase_date': clean_date(r.get('Purchase Date')),
            'purchase_price': clean_num(r.get('Purchase Price')),
            'maintenance_specs': clean_str(r.get('Maintenance Specs')),
            'maintenance_schedule': clean_str(r.get('Maintenance Schedule')),
            'last_maintenance_date': clean_date(r.get('Last Maintenance Date')),
            'next_maintenance_date': clean_date(r.get('Next Maintenance Date')),
            'photo_url': clean_str(r.get('Equipment Photo URL')),
            'notes': clean_str(r.get('Notes')),
        })
    batch_insert('equipment', equip_list)

    # ── 2n. Equipment Assignments ────────────────────────────────────────
    print('Importing equipment assignments...')
    ea_rows = read_sheet(wb, 'Equipment Assignment')
    equip_assigns = []
    for r in ea_rows:
        eq_code = clean_str(r.get('Equipment Code'))
        if not eq_code:
            continue
        eq_id = equipment_ids.get(eq_code)
        if not eq_id:
            continue

        staff_code = clean_str(r.get('Assigned Employee Code'))
        site_code = clean_str(r.get('Assigned Site Code'))

        equip_assigns.append({
            'id': gen_uuid(),
            'tenant_id': TENANT_ID,
            'equipment_id': eq_id,
            'staff_id': staff_ids.get(staff_code),
            'site_id': site_ids.get(site_code),
            'assigned_date': clean_date(r.get('Assignment Date')) or datetime.now().strftime('%Y-%m-%d'),
            'returned_date': clean_date(r.get('Return Date')),
            'notes': clean_str(r.get('Notes')),
        })
    batch_insert('equipment_assignments', equip_assigns)

    # ── 2o. Supply Assignments → site_supplies ───────────────────────────
    print('Importing site supplies (supply assignments)...')
    sa_rows = read_sheet(wb, 'Supply Assignment')
    site_supplies = []
    seen_ss = set()
    for r in sa_rows:
        site_code = clean_str(r.get('\U0001f3e2 Site_Code'))
        supply_code = clean_str(r.get('\U0001f3f7\ufe0f Supply_Code'))
        supply_name = clean_str(r.get('\U0001f4e6 Supply_Name'))
        if not site_code or not supply_code:
            continue
        site_id = site_ids.get(site_code)
        if not site_id:
            continue
        key = (site_id, supply_code)
        if key in seen_ss:
            continue
        seen_ss.add(key)

        site_supplies.append({
            'id': gen_uuid(),
            'tenant_id': TENANT_ID,
            'site_id': site_id,
            'name': supply_name or supply_code,
            'category': None,
            'notes': clean_str(r.get('Notes')),
        })
    batch_insert('site_supplies', site_supplies)

    # ── 2p. Inventory Counts ─────────────────────────────────────────────
    # NOTE: Excel headers are MISALIGNED with data. Actual data mapping:
    #   Col 0 (📊 Count ID)       → count_code (CNT-*)
    #   Col 1 (🔖 Count Code)     → site_code (SIT-*)
    #   Col 2 (🏢 Site Code)      → counted_by name
    #   Col 3 (📝 Form Code)      → count_date
    #   Col 6 (⏰ Count Timestamp) → notes
    print('Importing inventory counts...')
    ic_rows = read_sheet(wb, 'Inventory Count')
    counts = []
    for r in ic_rows:
        # Use ACTUAL data positions (headers are wrong)
        count_code = clean_str(r.get('\U0001f4ca Count ID'))  # col 0 = count code
        site_code = clean_str(r.get('\U0001f516 Count Code'))  # col 1 = site code
        if not count_code or not site_code:
            continue
        site_id = site_ids.get(site_code)
        if not site_id:
            continue

        cid = gen_uuid()
        count_ids[count_code] = cid

        # Col 3 (📝 Form Code) is actually count_date
        count_date = clean_date(r.get('\U0001f4dd Form Code'))
        if not count_date:
            count_date = datetime.now().strftime('%Y-%m-%d')

        # Col 6 (⏰ Count Timestamp) is actually notes
        notes = clean_str(r.get('\u23f0 Count Timestamp'))

        counts.append({
            'id': cid,
            'tenant_id': TENANT_ID,
            'count_code': count_code,
            'site_id': site_id,
            'count_date': count_date,
            'status': 'COMPLETED',
            'notes': notes,
        })
    batch_insert('inventory_counts', counts)

    # ── 2q. Inventory Count Details ──────────────────────────────────────
    # NOTE: Excel headers also misaligned. Actual data mapping:
    #   Col 0 (🔢 Detail ID)       → count_code (FK to parent)
    #   Col 2 (🏷️ Supply Code)     → supply description with [CODE] in brackets
    #   Col 3 (📦 Supply Category)  → quantity counted
    print('Importing inventory count details...')
    # Build supply name → supply_id map for fuzzy matching
    supply_name_map = {}
    for r in read_sheet(wb, 'Supply'):
        scode = clean_str(r.get('\U0001f3f7\ufe0f Supply_Code'))
        sname = clean_str(r.get('\U0001f1fa\U0001f1f8 Supply_Name_EN'))
        if scode and sname and scode in supply_ids:
            supply_name_map[sname.upper()] = supply_ids[scode]

    icd_rows = read_sheet(wb, 'Inventory Count Detail')
    details = []
    for r in icd_rows:
        # Col 0 = count_code (parent FK)
        parent_count_code = clean_str(r.get('\U0001f522 Detail ID'))
        if not parent_count_code:
            continue
        count_id = count_ids.get(parent_count_code)
        if not count_id:
            continue

        # Col 2 = supply description like "CLEANER FOO BAR [CODE]"
        supply_desc = clean_str(r.get('\U0001f3f7\ufe0f Supply Code'))
        if not supply_desc:
            continue

        # Try to match supply by name (strip bracketed code)
        name_part = re.sub(r'\s*\[.*?\]\s*$', '', supply_desc).strip().upper()
        supply_id = supply_name_map.get(name_part)
        if not supply_id:
            # Try partial match
            for sname, sid in supply_name_map.items():
                if sname.startswith(name_part[:30]) or name_part.startswith(sname[:30]):
                    supply_id = sid
                    break
        if not supply_id:
            continue

        # Col 3 = actual quantity
        qty = clean_num(r.get('\U0001f4e6 Supply Category'), 0)

        details.append({
            'id': gen_uuid(),
            'tenant_id': TENANT_ID,
            'count_id': count_id,
            'supply_id': supply_id,
            'actual_qty': qty,
            'notes': None,
        })
    batch_insert('inventory_count_details', details)

    # ── 2r. Update system_sequences ──────────────────────────────────────
    print('Updating system_sequences...')
    prefix_maxes = {}
    for code_map, prefix in [(client_ids, 'CLI'), (site_ids, 'SIT'), (supply_ids, 'SUP')]:
        for code in code_map.keys():
            parts = code.split('-')
            if len(parts) >= 2:
                try:
                    num = int(parts[1])
                    prefix_maxes[prefix] = max(prefix_maxes.get(prefix, 0), num)
                except ValueError:
                    pass

    for code in staff_ids.keys():
        # Skip suffixed variants — only process base codes (STF-NNNN, not STF-NNNN-A/B)
        if re.match(r'^STF-\d+-[AB]$', code):
            continue
        parts = code.split('-')
        if len(parts) >= 2:
            try:
                num = int(parts[1])
                prefix_maxes['STF'] = max(prefix_maxes.get('STF', 0), num)
            except ValueError:
                pass

    for code in job_ids.keys():
        parts = code.split('-')
        if len(parts) >= 2:
            try:
                num = int(parts[1])
                prefix_maxes['JOB'] = max(prefix_maxes.get('JOB', 0), num)
            except ValueError:
                pass

    for code in task_ids.keys():
        parts = code.split('-')
        if len(parts) >= 2:
            try:
                num = int(parts[1])
                prefix_maxes['TSK'] = max(prefix_maxes.get('TSK', 0), num)
            except ValueError:
                pass

    for code in service_ids.keys():
        parts = code.split('-')
        if len(parts) >= 2:
            try:
                num = int(parts[1])
                prefix_maxes['SER'] = max(prefix_maxes.get('SER', 0), num)
            except ValueError:
                pass

    for prefix, max_val in prefix_maxes.items():
        url = f'{SUPABASE_URL}/rest/v1/system_sequences?tenant_id=eq.{TENANT_ID}&prefix=eq.{prefix}'
        req = urllib.request.Request(url,
            data=json.dumps({'current_value': max_val}).encode(),
            headers={**HEADERS, 'Prefer': 'return=minimal'},
            method='PATCH')
        try:
            urllib.request.urlopen(req)
            print(f'  {prefix}: set to {max_val}')
        except urllib.error.HTTPError as e:
            print(f'  {prefix}: ERROR - {e.read().decode()[:100]}')

    # ── Summary ──────────────────────────────────────────────────────────
    print(f'\n=== IMPORT COMPLETE ===')
    print(f'Clients:              {len(client_ids)}')
    print(f'Sites:                {len(site_ids)}')
    print(f'Staff:                {len(staff_ids)}')
    print(f'Services:             {len(service_ids)}')
    print(f'Tasks:                {len(task_ids)}')
    print(f'Service Tasks:        {len(service_tasks)}')
    print(f'Site Jobs:            {len(job_ids)}')
    print(f'Job Tasks:            {len(job_tasks)}')
    print(f'Supplies:             {len(supply_ids)}')
    print(f'Equipment:            {len(equipment_ids)}')
    print(f'Equipment Assigns:    {len(equip_assigns)}')
    print(f'Site Supplies:        {len(site_supplies)}')
    print(f'Subcontractors:       {len(subcontractor_ids)}')
    print(f'Positions:            {len(position_ids)}')
    print(f'Inventory Counts:     {len(count_ids)}')
    print(f'Inventory Details:    {len(details)}')
    print(f'Lookups:              {len(lookups)}')


# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    delete_all_data()
    import_data()
    print('\nDone!')
